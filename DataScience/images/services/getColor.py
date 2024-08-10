from images.services.colorConvertService import ColorConvert
from images.services.selectcolorService import SelectColor
from images.services.RemoveWhiteMargin import CropMargin
from images.services.extract_color import Extract_Color
from images.services.blobService import BlobImages
from images.services.getImages import GetImages
from images.services.getImage import GetImage
import images.services.DominantColorService
from django.http import HttpResponse
from DataScience import settings
from io import BytesIO
import openpyxl
import json
import math



class GetColors:

    try:

        def get_colors(self,vendor):

            wb=openpyxl.Workbook()
            sheet=wb.active
            sheet.cell(row=1,column=1,value='SKU')
            sheet.cell(row=1,column=2,value='ImageURL')
            sheet.cell(row=1,column=3,value='HSV Color Code')
            sheet.cell(row=1,column=4,value='Dominant Color Name')

            first_response = GetImages().get_products(vendor,0)
            first_raw_data = first_response.content
            first_result = json.loads(first_raw_data)
            total=first_result['data']['total']
            bound=40
            count=math.ceil(total/bound)
            for i in range(231,count+1):
                response=GetImage().get_product(vendor,i)
                raw_data=response.content
                result=json.loads(raw_data)
                if result is not []:
                    data_information=result
                    increaser=0
                    for item in data_information:
                        if 'sku' in item:
                            sku = item['sku']
                            sheet.cell(row=(i*bound)+increaser+2,column=1,value=sku)
                            wb.save('static/excel/'+vendor+'.xlsx')
                            if 'images' in item:
                                if len(item['images'])>0:
                                    imageName=item['images'][0]['name']
                                    if 'imagePrefixPath' in item:
                                        imageprefixpath=item['imagePrefixPath']
                                        base_url = settings.AZURE_HTTP_HOST
                                        image_url = base_url + '/' + imageprefixpath + '/' + imageName + '_S.jpg'
                                        sheet.cell(row=(i*bound)+increaser+2,column=2,value=image_url)
                                        wb.save('static/excel/'+vendor+'.xlsx')
                                        id=1
                                        vendor=vendor
                                        collection=None
                                        design=None
                                        api='getImageColor'
                                        file=BlobImages.to_blob(id,image_url,imageName,vendor,collection,design,api)
                                        if file is not None:
                                            image_file=CropMargin().crop_margin(BytesIO(file))

                                            if image_file is not None:

                                                image_color=Extract_Color().ext_color(image_file)
                                                hsv_color = ColorConvert().convert_to_hsv(image_color)
                                                h=math.floor(hsv_color[0])
                                                s=math.floor(hsv_color[1])
                                                v=math.floor(hsv_color[2])
                                                hsv=(h,s,v)

                                                color_instance = SelectColor()
                                                dominant_color = color_instance.close_color(hsv_color)
                                                sheet.cell(row=(i * bound) + increaser + 2, column=3, value=str(hsv))
                                                sheet.cell(row=(i * bound) + increaser + 2, column=4, value=dominant_color)
                                                wb.save('static/excel/'+vendor+'.xlsx')
                                        else:

                                            sheet.cell(row=(i * bound) + increaser + 2, column=3, value='There is no image from the URL')
                                            wb.save('static/excel/'+vendor+'.xlsx')

                                        increaser += 1
                            else:

                                sheet.cell(row=(i*bound)+increaser+2, column=2, value='The Product has no image')
                                wb.save('static/excel/'+vendor+'.xlsx')
                                increaser += 1
                        else:

                            sheet.cell(row=(i*bound)+increaser+2, column=1, value='The product has no sku')
                            wb.save('static/excel/'+vendor+'.xlsx')
                            increaser += 1

                else:
                    GetColors.get_colors(self,vendor)


    except Exception as e:
        print('***** error in getting the color of images *****',str(e))

